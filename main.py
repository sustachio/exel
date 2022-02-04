from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter

# class for the whole thingy
class main:
  def __init__(self, wb):
    # get the excel file(s)
    self.wb = wb
    self.ws = wb.active
    
    self.ws.delete_rows(1, self.ws.max_row+1)
    self.fill((3,3),(10,10))

  # util functions 
  def makeCell(self, x, y, value):
    self.ws.cell(column=x, row=y, value=value)


  # fill sheet from 2 cords with a their cordiante
  def fill(self, start, end):
    for _column in range(end[0]):
      for _row in range(end[1]):
        
        # get the correct row that it is editing
        column = _column + start[0]
        row    = _row    + start[1]
      
        
        # put the cord of the cell into the cell (ex. B4, C9, AV39, ect.)
        cellValue = f"{get_column_letter(column)}{row}"
        self.makeCell(column, row, cellValue)
        
        
  def save(self):
    self.wb.save("sample.xlsx")

# grab the active worksheet
wb = load_workbook("excel_file1.xlsx")
main = main(wb)
main.save()